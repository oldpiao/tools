import pandas as pd
import os
from openpyxl import load_workbook


__all__ = [
    'excel_add_sheet', 'excel_update_sheet', 'excel_update_sheet2'
]


# enter code here
# dataframe: 需要写入excel的数据
# outfile：输出的文件地址
# name: sheet_name的文件名称

def excel_update_sheet(dataframe, outfile, sheet_name='Sheet1', header=False, index=None):
    """如果该页已存在，覆盖该页内容，其他页不变，默认是会将该页的名称后加数字，该方法将其改为修改"""
    writer = pd.ExcelWriter(outfile, enginge='openpyxl')
    if os.path.exists(outfile) is False:
        dataframe.to_excel(writer, sheet_name, header=header, index=index)
    else:
        book = load_workbook(writer.path)
        name_index = None
        if sheet_name in book.sheetnames:
            # book[name] = dataframe
            name_index = book.sheetnames.index(sheet_name)
            del book[sheet_name]
        writer.book = book
        dataframe.to_excel(excel_writer=writer, sheet_name=sheet_name, header=header, index=index)
        # a[:b] + [a[-1]] + a[b:-1]
        if name_index is not None:
            writer.book._sheets = writer.book._sheets[:name_index] + [writer.book._sheets[-1]] + writer.book._sheets[name_index:-1]
    writer.save()
    writer.close()


def excel_update_sheet2(dataframe, outfile, sheet_name='Sheet1', header=False, index=None):
    """覆盖写入，如果有第二第三...页会被删除掉"""
    dataframe.to_excel(excel_writer=outfile, sheet_name=sheet_name, header=header, index=index)


def excel_add_sheet(dataframe, outfile, sheet_name='Sheet1', header=False, index=None):
    """将处理结果转为excel中的新增页，如果该页已存在，会将该页的名称后加数字"""
    writer = pd.ExcelWriter(outfile, enginge='openpyxl')
    if os.path.exists(outfile) is False:
        dataframe.to_excel(writer, sheet_name, header=header, index=index)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        dataframe.to_excel(excel_writer=writer, sheet_name=sheet_name, header=header, index=index)
    writer.save()
    writer.close()


if __name__ == '__main__':
    filepath = 'D:/job/审计公告/审计问题分词/split_words/test/text.xlsx'
    # df = pd.read_excel(filepath, sheet_name='Sheet1', header=None)
    df = pd.read_excel(filepath, header=None)
    print(df)
    df[4] = 15
    print(df)
    excel_update_sheet(df, filepath, 'Sheet1')
    # excel_add_sheet(df, filepath, 'Sheet1')
